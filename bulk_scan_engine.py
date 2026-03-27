# """
# Bulk IP Scan Engine - Nmap, Nikto, parsing, and report generation.
# Cross-platform (Windows, Linux, macOS).
# """
# import subprocess
# import re
# import sys
# import shutil
# import platform
# from datetime import datetime

# # Directory for storing bulk scan reports
# REPORTS_DIR = 'reports'
# try:
#     import os
#     if not os.path.exists(REPORTS_DIR):
#         os.makedirs(REPORTS_DIR)
# except Exception:
#     pass


# def _is_windows():
#     return platform.system().lower() == 'windows'


# def _check_tool(name):
#     """Check if a tool is available on PATH. Cross-platform."""
#     return shutil.which(name) is not None


# def _ping_cmd(ip):
#     """Return ping command args for current OS."""
#     if _is_windows():
#         return ['ping', '-n', '1', '-w', '1000', ip]
#     return ['ping', '-c', '1', '-W', '1', ip]


# def log_output(message):
#     """Log output to console only."""
#     print(message)
#     sys.stdout.flush()


# def run_nmap_scan(ip, modules=None, port_depth='full', stopped_callback=None):
#     """
#     Run nmap scan on the target IP with configurable options.
#     modules: list of 'ports', 'services', 'os', 'vuln'
#     port_depth: 'quick' (top 100), 'standard' (top 500), 'deep' (top 1000), 'full' (ALL 1-65535)
#     For 'full': two-phase scan - first discover ALL open ports (fast), then get versions only on those.
#     """
#     modules = modules or ['ports', 'services']
#     try:
#         if not _check_tool('nmap'):
#             log_output(f"[NMAP] WARNING: Nmap not found! Using mock data for {ip}")
#             return generate_mock_nmap_output(ip)

#         # Port range
#         port_specs = {
#             'quick': ['--top-ports', '100'],
#             'standard': ['--top-ports', '500'],
#             'deep': ['--top-ports', '1000'],
#             'full': ['-p', '1-65535'],
#         }
#         port_args = port_specs.get(port_depth, port_specs['full'])

#         is_full_scan = (port_depth == 'full')

#         def _stopped():
#             return stopped_callback() if callable(stopped_callback) else False

#         if is_full_scan:
#             # PHASE 1: Port discovery ONLY - no service/OS/vuln (makes full scan feasible)
#             # Service detection on 65535 ports would take hours; we do it in phase 2 on found ports only
#             log_output(f"[NMAP] Phase 1: Discovering ALL open ports on {ip} (1-65535)...")
#             discovery_args = ['nmap', '-T5', '-Pn', '--open', '-sT', '--min-rate', '500',
#                              '-p', '1-65535', ip]
#             timeout = 2400  # 40 min for full port discovery
#             proc = subprocess.run(discovery_args, capture_output=True, text=True, timeout=timeout)
#             result = proc.stdout or ''
#             if not result.strip():
#                 return generate_mock_nmap_output(ip)

#             # Parse open ports from phase 1
#             phase1_data = parse_nmap_output(result)
#             open_ports = [p['port'] for p in phase1_data['ports']]
#             if not open_ports:
#                 log_output(f"[NMAP] No open ports found on {ip}")
#                 return result

#             log_output(f"[NMAP] Found {len(open_ports)} open port(s). Phase 2: Service detection...")
#             if _stopped():
#                 return result

#             # PHASE 2: Service detection ONLY on found ports (fast)
#             if 'services' in modules and open_ports:
#                 port_list = ','.join(open_ports)
#                 svc_args = ['nmap', '-T4', '-Pn', '-sV', '--version-intensity', '5',
#                             '-p', port_list, ip]
#                 if 'os' in modules:
#                     svc_args.extend(['-O', '--osscan-guess'])
#                 if 'vuln' in modules:
#                     svc_args.extend(['--script', 'vuln'])
#                 proc2 = subprocess.run(svc_args, capture_output=True, text=True, timeout=600)
#                 result = proc2.stdout or result
#             else:
#                 # Merge OS/vuln if requested (run on found ports)
#                 if ('os' in modules or 'vuln' in modules) and open_ports:
#                     port_list = ','.join(open_ports)
#                     extra_args = ['nmap', '-T4', '-Pn', '-p', port_list, ip]
#                     if 'os' in modules:
#                         extra_args.extend(['-O', '--osscan-guess'])
#                     if 'vuln' in modules:
#                         extra_args.extend(['--script', 'vuln'])
#                     proc2 = subprocess.run(extra_args, capture_output=True, text=True, timeout=300)
#                     if proc2.stdout:
#                         result = proc2.stdout

#             return result

#         # Non-full: single pass with all options
#         nmap_args = ['nmap', '-T4', '-Pn', '--open', '-sT'] + port_args
#         if 'services' in modules:
#             nmap_args.extend(['-sV', '--version-intensity', '5'])
#         if 'os' in modules:
#             nmap_args.extend(['-O', '--osscan-guess'])
#         if 'vuln' in modules:
#             nmap_args.extend(['--script', 'vuln'])

#         log_output(f"[NMAP] Scanning {ip} (modules: {', '.join(modules)}, ports: {port_depth})")
#         timeouts = {'quick': 60, 'standard': 90, 'deep': 180}
#         timeout = timeouts.get(port_depth, 180)
#         proc = subprocess.run(
#             nmap_args + [ip],
#             capture_output=True,
#             text=True,
#             timeout=timeout,
#         )
#         result = proc.stdout or ''
#         if proc.returncode != 0 and proc.stderr:
#             log_output(f"[NMAP] Stderr: {proc.stderr[:200]}")
#         if not result.strip():
#             return generate_mock_nmap_output(ip)
#         return result
#     except subprocess.TimeoutExpired:
#         log_output(f"[NMAP] Timeout for {ip}, using mock data")
#         return generate_mock_nmap_output(ip)
#     except FileNotFoundError:
#         log_output("[NMAP] Nmap not installed. Using mock data.")
#         return generate_mock_nmap_output(ip)
#     except Exception as e:
#         log_output(f"[NMAP] Error: {e}")
#         return generate_mock_nmap_output(ip)


# def run_nikto_scan(ip):
#     """Run nikto scan on the target IP (web vulnerability scan)."""
#     try:
#         if not _check_tool('nikto'):
#             log_output(f"[NIKTO] WARNING: Nikto not found! Using mock data for {ip}")
#             return generate_mock_nikto_output(ip)

#         log_output(f"[NIKTO] Scanning http://{ip}")
#         proc = subprocess.run(
#             ['nikto', '-h', f'http://{ip}', '-output', '-'],
#             capture_output=True,
#             text=True,
#             timeout=90,
#         )
#         result = proc.stdout or ''
#         if not result.strip():
#             return generate_mock_nikto_output(ip)
#         return result
#     except subprocess.TimeoutExpired:
#         log_output(f"[NIKTO] Timeout for {ip}")
#         return generate_mock_nikto_output(ip)
#     except FileNotFoundError:
#         log_output("[NIKTO] Nikto not installed. Using mock data.")
#         return generate_mock_nikto_output(ip)
#     except Exception as e:
#         log_output(f"[NIKTO] Error: {e}")
#         return generate_mock_nikto_output(ip)


# def generate_mock_nmap_output(ip):
#     """Generate mock nmap output when nmap is unavailable."""
#     return f"""
# Nmap scan report for {ip}
# Host is up (0.0010s latency).

# PORT     STATE SERVICE    VERSION
# 22/tcp   open  ssh        OpenSSH 8.2p1 Ubuntu 4ubuntu0.5 (Ubuntu Linux; protocol 2.0)
# 80/tcp   open  http       Apache httpd 2.4.41 ((Ubuntu))
# 443/tcp  open  ssl/http   nginx 1.18.0 (Ubuntu)
# 3306/tcp open  mysql      MySQL 5.7.30-0ubuntu0.18.04.1

# Service Info: OS: Linux; CPE: cpe:/o:linux:linux_kernel

# Host script results:
# | ssl-ccs-injection:
# |   VULNERABLE:
# |   SSL/TLS MITM vulnerability (CCS Injection)
# |     State: VULNERABLE
# |     References: CVE-2014-0224

# Nmap done: 1 IP address (1 host up) scanned in 25.42 seconds
# """


# def generate_mock_nikto_output(ip):
#     """Generate mock nikto output when nikto is unavailable."""
#     return f"""
# - Nikto v2.1.6
# + Target IP:          {ip}
# + Target Port:        80
# + Start Time:         {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
# + Server: Apache/2.4.41 (Ubuntu)
# + The anti-clickjacking X-Frame-Options header is not present.
# + The X-XSS-Protection header is not defined.
# + The X-Content-Type-Options header is not set.
# + Apache/2.4.41 appears to be outdated (current is at least Apache/2.4.54).
# + Allowed HTTP Methods: GET, POST, OPTIONS, HEAD
# + OSVDB-3268: /config/: Directory indexing found.
# + End Time:           {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
# """


# def parse_nmap_output(nmap_output):
#     """Parse nmap output to extract hostname, OS, ports, and vulnerabilities."""
#     data = {
#         'hostname': 'Unknown',
#         'os': 'Unknown',
#         'ports': [],
#         'vulnerabilities': []
#     }
#     if not nmap_output or len(nmap_output.strip()) < 10:
#         return data

#     # Hostname
#     m = re.search(r'Nmap scan report for (.+?)(?:\s*\(|$)', nmap_output)
#     if m:
#         data['hostname'] = m.group(1).strip()

#     # OS patterns
#     os_patterns = [
#         r'OS details:\s*(.+?)(?:\n|$)',
#         r'Running:\s*(.+?)(?:\n|$)',
#         r'OS:\s*(.+?)(?:;|\n|$)',
#         r'Aggressive OS guesses:\s*(.+?)(?:\(|,|\n)',
#         r'Service Info:\s*OS:\s*(.+?)(?:;|\n|$)',
#     ]
#     for pat in os_patterns:
#         om = re.search(pat, nmap_output, re.MULTILINE | re.IGNORECASE)
#         if om:
#             data['os'] = om.group(1).strip()[:100]
#             break

#     # Ports: robust parsing for nmap output formats
#     # Format 1: "22/tcp   open  ssh     OpenSSH 8.2p1 Ubuntu..."
#     # Format 2: "22/tcp   open  ssh"
#     # Format 3: "22/tcp open ssh OpenSSH 8.2"
#     seen_ports = set()
#     for line in nmap_output.splitlines():
#         line = line.strip()
#         # Skip script output lines (start with |)
#         if line.startswith('|'):
#             continue
#         # Match: PORT/proto  open  SERVICE  [version...]
#         m = re.match(r'(\d+)/(tcp|udp)\s+open\s+(\S+)\s*(.*)$', line)
#         if m:
#             port, proto, svc, rest = m.group(1), m.group(2), m.group(3), (m.group(4) or '').strip()
#             if port in seen_ports:
#                 continue
#             seen_ports.add(port)
#             version = rest[:80] if rest else 'Unknown'
#             data['ports'].append({
#                 'port': port,
#                 'protocol': proto,
#                 'service': svc,
#                 'version': version,
#                 'state': 'open'
#             })

#     # CVE extraction (CVE-YYYY-NNNNN+)
#     for cve in set(re.findall(r'(CVE-\d{4}-\d{4,})', nmap_output, re.IGNORECASE)):
#         data['vulnerabilities'].append({
#             'name': cve,
#             'description': f'Vulnerability {cve} detected',
#             'severity': 'High',
#             'source': 'Nmap',
#             'cve': cve.upper(),
#             'port': None
#         })

#     # Script vulnerabilities
#     vuln_patterns = [
#         (r'\|\s+VULNERABLE:\s*\n\|\s+(.+?)(?:\n\|(?!\s+)|$)', 'High'),
#         (r'ssl-cert:.+?VULNERABLE', 'Medium'),
#         (r'ssl-poodle:.+?VULNERABLE', 'High'),
#         (r'ssl-dh-params:.+?WEAK', 'Medium'),
#         (r'http-csrf:.+?Vulnerable', 'Medium'),
#         (r'http-vuln.+?VULNERABLE', 'High'),
#     ]
#     for pattern, severity in vuln_patterns:
#         for m in re.finditer(pattern, nmap_output, re.MULTILINE | re.DOTALL):
#             desc = m.group(0).strip()[:200].replace('|', '').replace('\n', ' ')
#             name = 'Security Issue'
#             nm = re.search(r'(\w+(?:-\w+)*)', desc)
#             if nm:
#                 name = nm.group(1)
#             data['vulnerabilities'].append({
#                 'name': name,
#                 'description': desc,
#                 'severity': severity,
#                 'source': 'Nmap Script',
#                 'cve': None,
#                 'port': None
#             })

#     return data


# def parse_nikto_output(nikto_output):
#     """Parse nikto output to extract vulnerabilities."""
#     vulnerabilities = []
#     if not nikto_output or len(nikto_output.strip()) < 10:
#         return vulnerabilities

#     skip_patterns = [
#         'target ip:', 'target hostname:', 'target port:', 'start time:', 'end time:',
#         'server:', 'requests:', '----------', 'nikto v', 'no web server found',
#         'testing:', 'retrieved', 'ssl info:'
#     ]

#     for line in nikto_output.split('\n'):
#         line = line.strip()
#         if not line.startswith('+ '):
#             continue
#         finding = line[2:].strip()
#         if len(finding) < 10:
#             continue
#         if any(skip in finding.lower() for skip in skip_patterns):
#             continue

#         severity = 'Low'
#         fl = finding.lower()
#         high_kw = ['vulnerability', 'exploit', 'injection', 'sql', 'authentication bypass',
#                    'remote code', 'arbitrary file', 'shell', 'xss', 'csrf', 'command injection',
#                    'lfi', 'rfi', 'directory traversal', 'path traversal', 'cve-']
#         med_kw = ['outdated', 'deprecated', 'disclosure', 'information leak', 'misconfiguration',
#                   'weak', 'insecure', 'cleartext', 'unencrypted', 'default', 'exposed']
#         if any(k in fl for k in high_kw):
#             severity = 'High'
#         elif any(k in fl for k in med_kw):
#             severity = 'Medium'

#         cve_match = re.search(r'(CVE-\d{4}-\d{4,})', finding, re.IGNORECASE)
#         vuln_name = cve_match.group(1) if cve_match else 'Web Vulnerability'
#         if 'header' in fl:
#             vuln_name = 'Missing Security Header'
#         elif 'directory' in fl and ('index' in fl or 'list' in fl):
#             vuln_name = 'Directory Listing'
#         elif 'outdated' in fl:
#             vuln_name = 'Outdated Software'
#         elif 'method' in fl:
#             vuln_name = 'Unsafe HTTP Methods'
#         elif 'ssl' in fl or 'tls' in fl:
#             vuln_name = 'SSL/TLS Configuration Issue'

#         vulnerabilities.append({
#             'source': 'Nikto',
#             'name': vuln_name,
#             'description': finding,
#             'severity': severity,
#             'cve': cve_match.group(1).upper() if cve_match else None,
#             'port': 80
#         })

#     return vulnerabilities


# def get_remediation_steps(vulnerability):
#     """Get remediation steps based on vulnerability type."""
#     db = {
#         'ssl-ccs-injection': {
#             'steps': [
#                 'Update OpenSSL to the latest stable version',
#                 'Disable SSLv3 and use only TLS 1.2 or higher',
#                 'Review and update SSL/TLS configuration',
#                 'Implement Perfect Forward Secrecy (PFS)'
#             ],
#             'commands': [
#                 'sudo apt update && sudo apt upgrade openssl',
#                 'Edit /etc/ssl/openssl.cnf to disable SSLv3',
#                 'sudo systemctl restart apache2',
#                 'Test with: nmap --script ssl-enum-ciphers -p 443 <target>'
#             ]
#         },
#         'outdated_apache': {
#             'steps': [
#                 'Update Apache to the latest stable version',
#                 'Enable automatic security updates',
#                 'Review server configuration for security best practices',
#                 'Implement security headers'
#             ],
#             'commands': [
#                 'sudo apt update && sudo apt upgrade apache2',
#                 'sudo apt install unattended-upgrades',
#                 'Add security headers to /etc/apache2/conf-available/security.conf',
#                 'sudo systemctl restart apache2'
#             ]
#         },
#         'missing_headers': {
#             'steps': [
#                 'Configure X-Frame-Options header to prevent clickjacking',
#                 'Enable X-XSS-Protection header',
#                 'Set X-Content-Type-Options header',
#                 'Implement Content Security Policy (CSP)'
#             ],
#             'commands': [
#                 'Add to Apache config: Header set X-Frame-Options "SAMEORIGIN"',
#                 'Add: Header set X-XSS-Protection "1; mode=block"',
#                 'Add: Header set X-Content-Type-Options "nosniff"',
#                 'sudo systemctl reload apache2'
#             ]
#         },
#         'directory_indexing': {
#             'steps': [
#                 'Disable directory indexing in web server configuration',
#                 'Create index files for all directories',
#                 'Review file permissions',
#                 'Implement access controls'
#             ],
#             'commands': [
#                 'Edit httpd.conf or apache2.conf: Options -Indexes',
#                 'Create index.html files in all directories',
#                 'sudo chmod 750 /var/www/html/*',
#                 'sudo systemctl restart apache2'
#             ]
#         },
#         'default': {
#             'steps': [
#                 'Review the security advisory for the specific vulnerability',
#                 'Apply vendor-recommended patches immediately',
#                 'Test the fix in a staging environment',
#                 'Monitor for similar vulnerabilities'
#             ],
#             'commands': [
#                 'Check vendor security bulletins',
#                 'sudo apt update && sudo apt upgrade',
#                 'Review system logs for exploitation attempts',
#                 'Implement monitoring and alerting'
#             ]
#         }
#     }
#     desc = vulnerability.get('description', '').lower()
#     if 'ssl' in desc or 'tls' in desc:
#         return db['ssl-ccs-injection']
#     if 'apache' in desc and 'outdated' in desc:
#         return db['outdated_apache']
#     if 'x-frame-options' in desc or 'x-xss-protection' in desc or 'header' in desc:
#         return db['missing_headers']
#     if 'directory indexing' in desc or 'directory listing' in desc:
#         return db['directory_indexing']
#     return db['default']


# def scan_single_ip(ip, modules=None, stopped_callback=None, port_depth='full'):
#     """
#     Perform full scan on a single IP.
#     Returns dict with: ip, status ('online'|'offline'), hostname, os, ports, vulnerabilities, severity, scanned_at
#     stopped_callback: callable that returns True if scan should stop
#     """
#     modules = modules or ['ping', 'ports', 'services']
#     result = {
#         'ip': ip,
#         'status': 'offline',
#         'hostname': 'Unknown',
#         'os': 'Unknown',
#         'ports': [],
#         'open_ports': [],
#         'services': [],
#         'vulnerabilities': [],
#         'severity': 'none',
#         'risk': 'None',
#         'ping': False,
#         'scanned_at': datetime.now().isoformat(),
#     }

#     try:
#         # Ping
#         if 'ping' in modules:
#             ping_ok = subprocess.run(
#                 _ping_cmd(ip),
#                 capture_output=True,
#                 timeout=5
#             ).returncode == 0
#             result['ping'] = ping_ok
#         else:
#             ping_ok = True  # Assume reachable if we skip ping

#         # Nmap
#         nmap_modules = []
#         if 'ports' in modules or 'services' in modules:
#             nmap_modules.append('ports')
#         if 'services' in modules:
#             nmap_modules.append('services')
#         if 'os' in modules:
#             nmap_modules.append('os')
#         if 'vuln' in modules:
#             nmap_modules.append('vuln')
#         if not nmap_modules:
#             nmap_modules = ['ports', 'services']

#         if stopped_callback and stopped_callback():
#             return result

#         nmap_output = run_nmap_scan(ip, nmap_modules, port_depth=port_depth, stopped_callback=stopped_callback)
#         nmap_data = parse_nmap_output(nmap_output)

#         result['hostname'] = nmap_data['hostname']
#         result['os'] = nmap_data['os']
#         result['ports'] = nmap_data['ports']
#         result['open_ports'] = [str(p['port']) for p in nmap_data['ports']]
#         result['services'] = [f"{p['port']}/{p['service']}" for p in nmap_data['ports']]
#         result['vulnerabilities'] = list(nmap_data['vulnerabilities'])

#         # Nikto (only if web ports open)
#         if 'nikto' in modules:
#             has_web = any(
#                 p['port'] in ['80', '443', '8080', '8443']
#                 for p in nmap_data['ports']
#             )
#             if has_web:
#                 if stopped_callback and stopped_callback():
#                     pass
#                 else:
#                     nikto_output = run_nikto_scan(ip)
#                     nikto_vulns = parse_nikto_output(nikto_output)
#                     result['vulnerabilities'].extend(nikto_vulns)

#         # Severity / Risk
#         vulns = result['vulnerabilities']
#         high_sev = sum(1 for v in vulns if v.get('severity') == 'High')
#         crit_sev = sum(1 for v in vulns if v.get('severity') == 'Critical')
#         risky_ports = {'21', '23', '3389', '445', '3306'}
#         has_risky = any(p in risky_ports for p in result['open_ports'])

#         if crit_sev or (high_sev >= 2) or has_risky:
#             result['severity'] = 'critical'
#             result['risk'] = 'High'
#         elif high_sev or has_risky:
#             result['severity'] = 'high'
#             result['risk'] = 'High'
#         elif vulns:
#             result['severity'] = 'medium'
#             result['risk'] = 'Medium'
#         elif result['open_ports']:
#             result['severity'] = 'low'
#             result['risk'] = 'Medium'
#         elif ping_ok:
#             result['severity'] = 'none'
#             result['risk'] = 'Low'
#         else:
#             result['severity'] = 'none'
#             result['risk'] = 'None'

#         # Online if we got any useful data
#         if result['ports'] or ping_ok:
#             result['status'] = 'online'
#         else:
#             result['status'] = 'offline'

#     except Exception as e:
#         result['status'] = 'error'
#         result['error'] = str(e)
#         result['severity'] = 'none'

#     return result


# def create_bulk_excel_report(scan_results, filepath=None):
#     """Create Excel report for bulk scan results. Returns filepath."""
#     from openpyxl import Workbook
#     from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
#     from openpyxl.utils import get_column_letter

#     wb = Workbook()
#     wb.remove(wb.active)

#     # Styles
#     title_font = Font(name='Arial', size=20, bold=True, color='1F4E78')
#     header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
#     header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
#     border_thin = Border(
#         left=Side(style='thin', color='CCCCCC'),
#         right=Side(style='thin', color='CCCCCC'),
#         top=Side(style='thin', color='CCCCCC'),
#         bottom=Side(style='thin', color='CCCCCC')
#     )

#     # Summary sheet
#     ws = wb.create_sheet('Summary')
#     ws['A1'] = 'Bulk IP Security Scan Report'
#     ws['A1'].font = title_font
#     ws.merge_cells('A1:F1')
#     ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
#     ws['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
#     ws.merge_cells('A2:F2')

#     headers = ['IP Address', 'Hostname', 'OS', 'Open Ports', 'Vulnerabilities', 'Severity']
#     for col, h in enumerate(headers, 1):
#         c = ws.cell(row=4, column=col)
#         c.value = h
#         c.font = header_font
#         c.fill = header_fill
#         c.border = border_thin
#         c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

#     for idx, r in enumerate(scan_results):
#         row = 5 + idx
#         fill = PatternFill(start_color='F2F2F2' if idx % 2 == 0 else 'FFFFFF', end_color='F2F2F2' if idx % 2 == 0 else 'FFFFFF', fill_type='solid')
#         ws.cell(row=row, column=1, value=r.get('ip', '')).border = border_thin
#         ws.cell(row=row, column=1).fill = fill
#         ws.cell(row=row, column=2, value=r.get('hostname', 'Unknown')).border = border_thin
#         ws.cell(row=row, column=2).fill = fill
#         ws.cell(row=row, column=3, value=r.get('os', 'Unknown')).border = border_thin
#         ws.cell(row=row, column=3).fill = fill
#         ws.cell(row=row, column=4, value=len(r.get('ports', []))).border = border_thin
#         ws.cell(row=row, column=4).fill = fill
#         ws.cell(row=row, column=5, value=len(r.get('vulnerabilities', []))).border = border_thin
#         ws.cell(row=row, column=5).fill = fill
#         sev_cell = ws.cell(row=row, column=6, value=(r.get('severity') or 'none').upper())
#         sev_cell.border = border_thin
#         sev_cell.fill = fill
#         sev_cell.alignment = Alignment(horizontal='center')

#     for i in range(1, 7):
#         ws.column_dimensions[get_column_letter(i)].width = [18, 30, 35, 12, 15, 12][i - 1]

#     # Vulnerabilities sheet
#     wv = wb.create_sheet('Vulnerabilities')
#     wv['A1'] = 'Detected Vulnerabilities'
#     wv['A1'].font = Font(name='Arial', size=14, bold=True, color='C00000')
#     wv.merge_cells('A1:E1')
#     v_headers = ['IP Address', 'Vulnerability', 'Severity', 'Source', 'Description']
#     for col, h in enumerate(v_headers, 1):
#         c = wv.cell(row=3, column=col)
#         c.value = h
#         c.font = header_font
#         c.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
#         c.border = border_thin
#         c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

#     row = 4
#     for r in scan_results:
#         vulns = r.get('vulnerabilities', [])
#         if not vulns:
#             wv.cell(row=row, column=1, value=r.get('ip', '')).border = border_thin
#             wv.cell(row=row, column=2, value='No vulnerabilities detected').border = border_thin
#             wv.merge_cells(f'B{row}:E{row}')
#             row += 1
#         else:
#             for v in vulns:
#                 wv.cell(row=row, column=1, value=r.get('ip', '')).border = border_thin
#                 wv.cell(row=row, column=2, value=v.get('name', 'Unknown')).border = border_thin
#                 wv.cell(row=row, column=3, value=v.get('severity', 'Medium')).border = border_thin
#                 wv.cell(row=row, column=4, value=v.get('source', 'Nmap')).border = border_thin
#                 wv.cell(row=row, column=5, value=v.get('description', '')[:500]).border = border_thin
#                 wv.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
#                 row += 1

#     for i in range(1, 6):
#         wv.column_dimensions[get_column_letter(i)].width = [18, 30, 12, 12, 70][i - 1]

#     if not filepath:
#         filepath = f'{REPORTS_DIR}/bulk_scan_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
#     wb.save(filepath)
#     return filepath



"""
Bulk IP Scan Engine - Nmap, Nikto, parsing, and report generation.
Cross-platform (Windows, Linux, macOS).
"""
import os
import subprocess
import re
import sys
import shutil
import platform
from datetime import datetime

# Software version detection & outdated-software CVE enrichment
try:
    from software_version_detector import enrich_scan_result_with_software
    _SOFTWARE_DETECTION_ENABLED = True
except ImportError:
    _SOFTWARE_DETECTION_ENABLED = False
    def enrich_scan_result_with_software(result):
        return result

# OpenVAS / multi-layer OS detection (Ubuntu, Windows, Red Hat, etc.)
try:
    from openvas_os_detector import enrich_result_with_os, check_openvas_available
    _OPENVAS_OS_ENABLED = True
    # Read credentials from env BEFORE calling check (not after), so the check
    # uses the real password from .env, not the hardcoded 'admin' default.
    _ova_check = check_openvas_available(
        gvm_socket   = os.environ.get('OPENVAS_SOCKET', '').strip() or None,
        gvm_host     = os.environ.get('OPENVAS_HOST',     '127.0.0.1'),
        gvm_port     = int(os.environ.get('OPENVAS_PORT', '9390')),
        gvm_user     = os.environ.get('OPENVAS_USER',     'admin'),
        gvm_password = os.environ.get('OPENVAS_PASSWORD', 'admin'),
    )
    if _ova_check['available']:
        print(f"[OPENVAS] Connected via {_ova_check['method']} (GVM {_ova_check['version']})", flush=True)
    else:
        print(f"[OPENVAS] Not available — {_ova_check['error']}", flush=True)
        print(f"[OPENVAS] OS detection will use nmap/banner fallbacks", flush=True)
except ImportError:
    _OPENVAS_OS_ENABLED = False
    def enrich_result_with_os(result, openvas_config=None):
        return result

# OpenVAS connection config — reads from environment variables so you never
# hard-code credentials. Set these in your .env file:
#   OPENVAS_SOCKET=/run/gvmd/gvmd.sock   (preferred, same machine)
#   OPENVAS_HOST=127.0.0.1
#   OPENVAS_PORT=9390
#   OPENVAS_USER=admin
#   OPENVAS_PASSWORD=admin

def _auto_detect_gvm_socket():
    """
    Find the GVM socket file by checking all known locations.
    Respects OPENVAS_SOCKET env var; falls back to common WSL/Linux paths.
    Returns socket path string or '' if none found.
    """
    env_sock = os.environ.get('OPENVAS_SOCKET', '').strip()
    if env_sock and os.path.exists(env_sock):
        return env_sock
    candidates = [
        '/run/gvmd/gvmd.sock',          # Kali / Debian / Ubuntu default
        '/var/run/gvmd/gvmd.sock',      # older distros
        '/run/gvm/gvmd.sock',           # some community builds
        '/var/run/gvm/gvmd.sock',
        '/tmp/gvm/gvmd.sock',
        '/usr/local/var/run/gvmd.sock', # macOS homebrew
    ]
    for path in candidates:
        if os.path.exists(path):
            log_output(f"[OPENVAS] Auto-detected GVM socket: {path}")
            return path
    return ''

_OPENVAS_CONFIG = {
    'socket':   _auto_detect_gvm_socket(),
    'host':     os.environ.get('OPENVAS_HOST',     '127.0.0.1'),
    'port':     int(os.environ.get('OPENVAS_PORT', '9390')),
    'user':     os.environ.get('OPENVAS_USER',     'admin'),
    'password': os.environ.get('OPENVAS_PASSWORD', 'admin'),
}

# Directory for storing bulk scan reports
REPORTS_DIR = 'reports'
try:
    import os
    if not os.path.exists(REPORTS_DIR):
        os.makedirs(REPORTS_DIR)
except Exception:
    pass


def _is_windows():
    return platform.system().lower() == 'windows'


def _check_tool(name):
    """Check if a tool is available on PATH. Cross-platform."""
    return shutil.which(name) is not None


def _has_root_for_nmap():
    """OS detection (-O) requires root on Linux. Returns True if we can use it."""
    if _is_windows():
        return True  # Windows nmap doesn't use raw sockets for -O the same way
    try:
        return os.geteuid() == 0
    except (AttributeError, OSError):
        return False


def _ping_cmd(ip):
    """Return ping command args for current OS."""
    if _is_windows():
        return ['ping', '-n', '4', '-w', '1000', ip]
    return ['ping', '-c', '4', '-W', '2', ip]


def run_ping_detailed(ip):
    """
    Run ping and return a detailed dict with latency, RTT stats, packet loss, etc.
    Returns: {
        'alive': bool,
        'latency': float|None,   # avg RTT in ms
        'min_rtt': str,
        'max_rtt': str,
        'avg_rtt': str,
        'packets_sent': int,
        'packets_recv': int,
        'packet_loss': str,      # e.g. "0%"
        'raw': str               # full ping output
    }
    """
    detail = {
        'alive': False,
        'latency': None,
        'min_rtt': None,
        'max_rtt': None,
        'avg_rtt': None,
        'packets_sent': None,
        'packets_recv': None,
        'packet_loss': None,
        'raw': '',
    }
    try:
        proc = subprocess.run(
            _ping_cmd(ip),
            capture_output=True,
            text=True,
            timeout=10,
        )
        output = (proc.stdout or '') + (proc.stderr or '')
        detail['raw'] = output
        detail['alive'] = proc.returncode == 0

        if _is_windows():
            # Windows: "Minimum = 1ms, Maximum = 2ms, Average = 1ms"
            m = re.search(r'Minimum\s*=\s*(\d+)ms,\s*Maximum\s*=\s*(\d+)ms,\s*Average\s*=\s*(\d+)ms', output, re.I)
            if m:
                detail['min_rtt'] = m.group(1) + ' ms'
                detail['max_rtt'] = m.group(2) + ' ms'
                detail['avg_rtt'] = m.group(3) + ' ms'
                detail['latency'] = float(m.group(3))
            # Windows: "Packets: Sent = 4, Received = 4, Lost = 0 (0% loss)"
            pm = re.search(r'Sent\s*=\s*(\d+),\s*Received\s*=\s*(\d+),\s*Lost\s*=\s*\d+\s*\((\d+)%\s*loss\)', output, re.I)
            if pm:
                detail['packets_sent'] = int(pm.group(1))
                detail['packets_recv'] = int(pm.group(2))
                detail['packet_loss'] = pm.group(3) + '%'
        else:
            # Linux/Mac: "rtt min/avg/max/mdev = 0.123/0.456/0.789/0.100 ms"
            m = re.search(r'rtt\s+min/avg/max/\w+\s*=\s*([\d.]+)/([\d.]+)/([\d.]+)', output, re.I)
            if m:
                detail['min_rtt'] = m.group(1) + ' ms'
                detail['avg_rtt'] = m.group(2) + ' ms'
                detail['max_rtt'] = m.group(3) + ' ms'
                detail['latency'] = float(m.group(2))
            # Linux: "4 packets transmitted, 4 received, 0% packet loss"
            pm = re.search(r'(\d+)\s+packets\s+transmitted,\s*(\d+)\s+received,\s*([\d.]+)%\s+packet\s+loss', output, re.I)
            if pm:
                detail['packets_sent'] = int(pm.group(1))
                detail['packets_recv'] = int(pm.group(2))
                detail['packet_loss'] = pm.group(3) + '%'
            # If host alive but no RTT stats (ICMP blocked but host replied via other means)
            if detail['alive'] and not detail['latency']:
                # Try single-line time= extraction: "time=1.23 ms"
                tm = re.search(r'time[<=]([\d.]+)\s*ms', output, re.I)
                if tm:
                    val = float(tm.group(1))
                    detail['latency'] = val
                    detail['avg_rtt'] = str(val) + ' ms'
    except (subprocess.TimeoutExpired, OSError, Exception):
        pass
    return detail


def log_output(message):
    """Log output to console only."""
    print(message)
    sys.stdout.flush()


def run_nmap_discovery(ip):
    """
    Run a fast nmap host-discovery scan (no port scan) to get:
      - MAC address (works on LAN via ARP)
      - Hostname via reverse DNS (-R) and NetBIOS/SMB scripts
      - Latency from the 'Host is up (Xs latency)' line

    Uses 'nmap -sn -R' WITHOUT -Pn so ARP is used on local subnets.
    Falls back to empty dict on failure.
    Returns: { 'mac_address': str|None, 'hostname': str|None, 'latency_ms': float|None }
    """
    if not _check_tool('nmap'):
        return {}
    try:
        # -sn  = ping scan only (no port scan), allows ARP on LAN
        # -R   = always do reverse DNS resolution
        # --script nbstat = NetBIOS name lookup (Windows hostnames)
        # No -Pn so nmap uses ARP for local subnet hosts → gets MAC
        args = ['nmap', '-sn', '-R', '--script', 'nbstat', '--script-timeout', '5s', ip]
        proc = subprocess.run(args, capture_output=True, text=True, timeout=30)
        output = proc.stdout or ''
        log_output(f"[DISCOVERY] nmap -sn output for {ip}:\n{output[:400]}")

        result = {}

        # MAC address: "MAC Address: AA:BB:CC:DD:EE:FF (Vendor Name)"
        mac_m = re.search(r'MAC Address:\s*([0-9A-Fa-f]{2}(?::[0-9A-Fa-f]{2}){5})\s*(?:\(([^)]+)\))?', output)
        if mac_m:
            mac = mac_m.group(1).upper()
            vendor = mac_m.group(2)
            result['mac_address'] = f"{mac} ({vendor})" if vendor else mac

        # Hostname: "Nmap scan report for HOSTNAME (IP)" — hostname is before the paren
        # or "Nmap scan report for IP" — no hostname
        hn_m = re.search(r'Nmap scan report for\s+(.+?)\s*\([\d.]+\)', output)
        if hn_m:
            hostname = hn_m.group(1).strip()
            if hostname and not re.match(r'^\d+\.\d+\.\d+\.\d+$', hostname):
                result['hostname'] = hostname
        # Also try nbstat script output: "NetBIOS name: DESKTOP-XXXXX"
        if not result.get('hostname'):
            nb_m = re.search(r'NetBIOS name:\s*(\S+)', output, re.IGNORECASE)
            if nb_m:
                result['hostname'] = nb_m.group(1).strip().rstrip('.')

        # Latency from "Host is up (0.00050s latency)"
        lat_m = re.search(r'Host is up \(([\d.]+)s latency\)', output)
        if lat_m:
            result['latency_ms'] = round(float(lat_m.group(1)) * 1000, 2)

        return result
    except Exception as e:
        log_output(f"[DISCOVERY] nmap -sn failed for {ip}: {e}")
        return {}


def run_masscan_scan(ip, port_range='1-65535'):
    """Run masscan for fast port discovery. Requires root. Returns list of open port strings or []."""
    if not _check_tool('masscan'):
        return []
    if not _has_root_for_nmap():
        return []
    try:
        log_output(f"[MASSCAN] Phase 1: Scanning {ip} ({port_range})...")
        # masscan: -p1-65535, --rate=1000, output: open tcp PORT IP timestamp
        proc = subprocess.run(
            ['masscan', ip, '-p' + port_range, '--rate', '1000'],
            capture_output=True,
            text=True,
            timeout=300,
        )
        ports = []
        for line in (proc.stdout or '').splitlines():
            # Format: "open tcp 3389 192.168.1.1 1234567890"
            m = re.match(r'open\s+tcp\s+(\d+)\s+', line)
            if m:
                ports.append(m.group(1))
        ports = list(dict.fromkeys(ports))  # dedupe, preserve order
        if not ports and (proc.stderr or proc.returncode != 0):
            err = (proc.stderr or '')[:300].strip()
            log_output(f"[MASSCAN] No ports (rc={proc.returncode}): {err or 'no output'}")
        return ports
    except (subprocess.TimeoutExpired, FileNotFoundError, Exception) as e:
        log_output(f"[MASSCAN] Skipped: {e}")
        return []


def run_nuclei_scan(ip, web_ports=None):
    """Run nuclei template-based vuln scan on web URLs. Returns parsed vulnerabilities."""
    if not _check_tool('nuclei'):
        return []
    web_ports = web_ports or ['80', '443']
    urls = []
    if '80' in web_ports:
        urls.append(f'http://{ip}')
    if '443' in web_ports:
        urls.append(f'https://{ip}')
    if '8080' in web_ports:
        urls.append(f'http://{ip}:8080')
    if '8443' in web_ports:
        urls.append(f'https://{ip}:8443')
    if not urls:
        return []
    try:
        args = ['nuclei', '-silent', '-jsonl', '-no-color']
        for u in urls:
            args.extend(['-u', u])
        proc = subprocess.run(args, capture_output=True, text=True, timeout=120)
        return parse_nuclei_output(proc.stdout or '')
    except (subprocess.TimeoutExpired, FileNotFoundError, Exception) as e:
        log_output(f"[NUCLEI] Skipped: {e}")
        return []


def parse_nuclei_output(jsonl_output):
    """Parse nuclei JSONL output into vulnerability dicts."""
    import json
    vulns = []
    for line in (jsonl_output or '').strip().splitlines():
        if not line.strip():
            continue
        try:
            j = json.loads(line)
            info = j.get('info', {})
            name = info.get('name', 'Nuclei Finding')
            severity = (info.get('severity') or 'info').lower()
            severity = {'critical': 'Critical', 'high': 'High', 'medium': 'Medium', 'low': 'Low'}.get(severity, 'Low')
            desc = info.get('description', name) or name
            cve = None
            ref = info.get('reference')
            if isinstance(ref, list) and ref:
                for r in ref:
                    s = str(r) if isinstance(r, str) else str(r.get('url', r))
                    m = re.search(r'(CVE-\d{4}-\d{4,})', s, re.I)
                    if m:
                        cve = m.group(1).upper()
                        break
            if not cve:
                m = re.search(r'(CVE-\d{4}-\d{4,})', str(j), re.I)
                if m:
                    cve = m.group(1).upper()
            vulns.append({
                'name': name[:100],
                'description': (desc or name)[:200],
                'severity': severity,
                'source': 'Nuclei',
                'cve': cve,
                'port': None
            })
        except (json.JSONDecodeError, KeyError, TypeError):
            continue
    return vulns


def run_nmap_scan(ip, modules=None, port_depth='full', stopped_callback=None):
    """
    Run nmap scan on the target IP with configurable options.
    modules: list of 'ports', 'services', 'os', 'vuln'
    port_depth: 'quick' (top 100), 'standard' (top 500), 'deep' (top 1000), 'full' (ALL 1-65535)
    For 'full': two-phase scan - first discover ALL open ports (fast), then get versions only on those.
    """
    modules = modules or ['ports', 'services']
    try:
        if not _check_tool('nmap'):
            log_output(f"[NMAP] WARNING: Nmap not found! Using mock data for {ip}")
            return generate_mock_nmap_output(ip)

        # Port range
        port_specs = {
            'quick': ['--top-ports', '100'],
            'standard': ['--top-ports', '500'],
            'deep': ['--top-ports', '1000'],
            'full': ['-p', '1-65535'],
        }
        port_args = port_specs.get(port_depth, port_specs['full'])

        is_full_scan = (port_depth == 'full')

        def _stopped():
            return stopped_callback() if callable(stopped_callback) else False

        if is_full_scan:
            # PHASE 1: Fast port discovery - use Masscan if available (10x faster), else Nmap
            open_ports = run_masscan_scan(ip, '1-65535')
            if open_ports:
                log_output(f"[MASSCAN] Found {len(open_ports)} open port(s) on {ip}")
            elif _check_tool('masscan') and not _has_root_for_nmap():
                log_output("[MASSCAN] Skipped (requires root). Using Nmap - full scan may take 15-40 min.")
            if not open_ports:
                log_output(f"[NMAP] Phase 1: Discovering ALL open ports on {ip} (1-65535)...")
                # No -Pn: allow ARP on LAN so MAC address is captured
                # -R: reverse DNS resolution for hostname
                discovery_args = ['nmap', '-T5', '--open', '-sT', '-R',
                                 '--min-rate', '500', '-p', '1-65535', ip]
                timeout = 2400  # 40 min for full port discovery
                proc = subprocess.run(discovery_args, capture_output=True, text=True, timeout=timeout)
                result = proc.stdout or ''
                if not result.strip():
                    return generate_mock_nmap_output(ip)
                phase1_data = parse_nmap_output(result)
                open_ports = [p['port'] for p in phase1_data['ports']]
                result = result  # keep for phase 2 merge
            else:
                result = f"Nmap scan report for {ip}\nHost is up.\n"  # placeholder for merge
            if not open_ports:
                log_output(f"[NMAP] No open ports found on {ip}")
                return result

            log_output(f"[NMAP] Found {len(open_ports)} open port(s). Phase 2: Service detection...")
            if _stopped():
                return result

            # PHASE 2: Service/version detection on found ports only
            # -Pn kept here since we already know host is up from phase 1
            # -R for reverse DNS
            if 'services' in modules and open_ports:
                port_list = ','.join(open_ports)
                svc_args = ['nmap', '-T4', '-Pn', '-R', '-sV', '--version-intensity', '5',
                            '-sC',
                            '--script', 'ssh-hostkey,http-server-header,smb-os-discovery,nbstat,banner',
                            '-p', port_list, ip]
                if 'os' in modules and _has_root_for_nmap():
                    svc_args.extend(['-O', '--osscan-guess'])
                elif 'os' in modules:
                    log_output("[NMAP] Skipping OS fingerprint (-O) - requires root. OS inferred from service banners.")
                if 'vuln' in modules:
                    svc_args.extend(['--script', 'vuln'])
                proc2 = subprocess.run(svc_args, capture_output=True, text=True, timeout=600)
                result = proc2.stdout or result
            else:
                # Merge OS/vuln if requested (run on found ports)
                if ('os' in modules or 'vuln' in modules) and open_ports:
                    port_list = ','.join(open_ports)
                    extra_args = ['nmap', '-T4', '-Pn', '-R', '-p', port_list, ip]
                    if 'os' in modules and _has_root_for_nmap():
                        extra_args.extend(['-O', '--osscan-guess'])
                    elif 'os' in modules:
                        log_output("[NMAP] Skipping OS fingerprint (-O) - requires root.")
                    if 'vuln' in modules:
                        extra_args.extend(['--script', 'vuln'])
                    proc2 = subprocess.run(extra_args, capture_output=True, text=True, timeout=300)
                    if proc2.stdout:
                        result = proc2.stdout

            return result

        # Non-full: single pass with all options
        # No -Pn so ARP is used on LAN (gives MAC address)
        # -R for reverse DNS hostname resolution
        nmap_args = ['nmap', '-T4', '--open', '-sT', '-R'] + port_args
        if 'services' in modules:
            nmap_args.extend(['-sV', '--version-intensity', '5',
                              '-sC',
                              '--script', 'ssh-hostkey,http-server-header,smb-os-discovery,nbstat,banner'])
        if 'os' in modules and _has_root_for_nmap():
            nmap_args.extend(['-O', '--osscan-guess'])
        elif 'os' in modules:
            log_output("[NMAP] Skipping OS fingerprint (-O) - requires root. OS inferred from service banners.")
        if 'vuln' in modules:
            nmap_args.extend(['--script', 'vuln'])

        log_output(f"[NMAP] Scanning {ip} (modules: {', '.join(modules)}, ports: {port_depth})")
        timeouts = {'quick': 120, 'standard': 300, 'deep': 600}  # 2/5/10 min
        timeout = timeouts.get(port_depth, 600)
        proc = subprocess.run(
            nmap_args + [ip],
            capture_output=True,
            text=True,
            timeout=timeout,
        )
        result = proc.stdout or ''
        if proc.returncode != 0 and proc.stderr:
            log_output(f"[NMAP] Stderr: {proc.stderr[:200]}")
        if not result.strip():
            return generate_mock_nmap_output(ip)
        return result
    except subprocess.TimeoutExpired:
        log_output(f"[NMAP] Timeout for {ip}, using mock data")
        return generate_mock_nmap_output(ip)
    except FileNotFoundError:
        log_output("[NMAP] Nmap not installed. Using mock data.")
        return generate_mock_nmap_output(ip)
    except Exception as e:
        log_output(f"[NMAP] Error: {e}")
        return generate_mock_nmap_output(ip)


def run_nikto_scan(ip, web_ports=None):
    """Run nikto scan on the target IP. Scans both HTTP and HTTPS when ports 80/443 open.
    web_ports: list of open web ports e.g. ['80','443','8080','8443'] - if None, scans http only."""
    try:
        if not _check_tool('nikto'):
            log_output(f"[NIKTO] WARNING: Nikto not found! Using mock data for {ip}")
            return generate_mock_nikto_output(ip)

        web_ports = web_ports or ['80']
        urls = []
        if '80' in web_ports:
            urls.append(f'http://{ip}')
        if '443' in web_ports:
            urls.append(f'https://{ip}')
        if '8080' in web_ports:
            urls.append(f'http://{ip}:8080')
        if '8443' in web_ports:
            urls.append(f'https://{ip}:8443')
        if not urls:
            urls = [f'http://{ip}']

        all_output = []
        for url in urls:
            log_output(f"[NIKTO] Scanning {url}")
            args = ['nikto', '-h', url, '-output', '-']
            if url.startswith('https'):
                args.append('-nossl')  # Skip SSL cert verification for self-signed
            proc = subprocess.run(
                args,
                capture_output=True,
                text=True,
                timeout=90,
            )
            out = proc.stdout or ''
            if out.strip():
                all_output.append(out)
        result = '\n---\n'.join(all_output) if all_output else ''
        if not result.strip():
            return generate_mock_nikto_output(ip)
        return result
    except subprocess.TimeoutExpired:
        log_output(f"[NIKTO] Timeout for {ip}")
        return generate_mock_nikto_output(ip)
    except FileNotFoundError:
        log_output("[NIKTO] Nikto not installed. Using mock data.")
        return generate_mock_nikto_output(ip)
    except Exception as e:
        log_output(f"[NIKTO] Error: {e}")
        return generate_mock_nikto_output(ip)


def generate_mock_nmap_output(ip):
    """Generate mock nmap output when nmap is unavailable."""
    return f"""
Nmap scan report for {ip}
Host is up (0.0010s latency).

PORT     STATE SERVICE    VERSION
22/tcp   open  ssh        OpenSSH 8.2p1 Ubuntu 4ubuntu0.5 (Ubuntu Linux; protocol 2.0)
80/tcp   open  http       Apache httpd 2.4.41 ((Ubuntu))
443/tcp  open  ssl/http   nginx 1.18.0 (Ubuntu)
3306/tcp open  mysql      MySQL 5.7.30-0ubuntu0.18.04.1

Service Info: OS: Linux; CPE: cpe:/o:linux:linux_kernel

Host script results:
| ssl-ccs-injection:
|   VULNERABLE:
|   SSL/TLS MITM vulnerability (CCS Injection)
|     State: VULNERABLE
|     References: CVE-2014-0224

Nmap done: 1 IP address (1 host up) scanned in 25.42 seconds
"""


def generate_mock_nikto_output(ip):
    """Generate mock nikto output when nikto is unavailable."""
    return f"""
- Nikto v2.1.6
+ Target IP:          {ip}
+ Target Port:        80
+ Start Time:         {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
+ Server: Apache/2.4.41 (Ubuntu)
+ The anti-clickjacking X-Frame-Options header is not present.
+ The X-XSS-Protection header is not defined.
+ The X-Content-Type-Options header is not set.
+ Apache/2.4.41 appears to be outdated (current is at least Apache/2.4.54).
+ Allowed HTTP Methods: GET, POST, OPTIONS, HEAD
+ OSVDB-3268: /config/: Directory indexing found.
+ End Time:           {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""


def parse_nmap_output(nmap_output):
    """Parse nmap output to extract hostname, OS, ports, scan stats, and vulnerabilities."""
    data = {
        'hostname': 'Unknown',
        'os': 'Unknown',
        'ports': [],
        'vulnerabilities': [],
        # Scan metadata
        'mac_address': None,
        'scan_duration': None,
        'ports_scanned': None,
        'port_range': None,
        'scan_technique': 'TCP SYN',
    }
    if not nmap_output or len(nmap_output.strip()) < 10:
        return data

    # Hostname — two formats nmap uses:
    #   "Nmap scan report for HOSTNAME (192.168.1.x)"  ← has rDNS
    #   "Nmap scan report for 192.168.1.x"             ← no rDNS
    hn_paren = re.search(r'Nmap scan report for\s+(.+?)\s*\([\d.]+\)', nmap_output)
    if hn_paren:
        candidate = hn_paren.group(1).strip()
        if candidate and not re.match(r'^\d+\.\d+\.\d+\.\d+$', candidate):
            data['hostname'] = candidate
    else:
        # No parentheses — could be bare IP or hostname
        m = re.search(r'Nmap scan report for\s+(\S+)', nmap_output)
        if m:
            candidate = m.group(1).strip()
            if candidate and not re.match(r'^\d+\.\d+\.\d+\.\d+$', candidate):
                data['hostname'] = candidate
    # Sometimes: "Nmap scan report for 192.168.1.1\nHost is up"
    # And RDNS shows up in a different line

    # MAC address
    mac_m = re.search(r'MAC Address:\s*([0-9A-Fa-f:]{17})\s*(?:\(([^)]+)\))?', nmap_output)
    if mac_m:
        data['mac_address'] = mac_m.group(1)
        # If vendor available, append it
        if mac_m.group(2):
            data['mac_address'] = f"{mac_m.group(1)} ({mac_m.group(2)})"

    # OS patterns
    os_patterns = [
        r'OS details:\s*(.+?)(?:\n|$)',
        r'Running:\s*(.+?)(?:\n|$)',
        r'OS:\s*(.+?)(?:;|\n|$)',
        r'Aggressive OS guesses:\s*(.+?)(?:\(|,|\n)',
        r'Service Info:\s*OS:\s*(.+?)(?:;|\n|$)',
    ]
    for pat in os_patterns:
        om = re.search(pat, nmap_output, re.MULTILINE | re.IGNORECASE)
        if om:
            data['os'] = om.group(1).strip()[:100]
            break

    # Scan duration: "Nmap done: 1 IP address (1 host up) scanned in 25.42 seconds"
    dur_m = re.search(r'scanned in ([\d.]+) seconds', nmap_output, re.IGNORECASE)
    if dur_m:
        secs = float(dur_m.group(1))
        data['scan_duration'] = f"{secs:.1f}s"

    # Scan technique
    if '-sS' in nmap_output or 'SYN' in nmap_output:
        data['scan_technique'] = 'TCP SYN'
    elif '-sT' in nmap_output or 'CONNECT' in nmap_output.upper():
        data['scan_technique'] = 'TCP SYN'

    # Ports: robust parsing — also extract CPE and extra_info
    # nmap -sV output format:
    #   PORT      STATE  SERVICE        VERSION
    #   22/tcp    open   ssh            OpenSSH 8.2p1 Ubuntu 4ubuntu0.5
    # With -sC scripts, extra info may appear on next | lines
    # CPE appears as: "cpe:/a:openbsd:openssh:8.2p1"
    seen_ports = set()
    lines = nmap_output.splitlines()
    port_list_for_range = []

    for i, line in enumerate(lines):
        raw_line = line
        line = line.strip()
        # Skip script output lines (start with |)
        if line.startswith('|'):
            continue
        # Match: PORT/proto  open  SERVICE  [version...]
        m = re.match(r'(\d+)/(tcp|udp)\s+open\s+(\S+)\s*(.*)$', line)
        if m:
            port, proto, svc, rest = m.group(1), m.group(2), m.group(3), (m.group(4) or '').strip()
            if port in seen_ports:
                continue
            seen_ports.add(port)
            port_list_for_range.append(int(port))

            # Extract CPE from version string: "OpenSSH 8.2p1 ... cpe:/a:openbsd:openssh:8.2p1"
            cpe = None
            cpe_m = re.search(r'(cpe:/[^\s]+)', rest, re.IGNORECASE)
            if cpe_m:
                cpe = cpe_m.group(1)
                rest = rest[:cpe_m.start()].strip()

            # extra_info: text in parentheses at end of version line
            extra_info = None
            ei_m = re.search(r'\(([^)]+)\)\s*$', rest)
            if ei_m:
                extra_info = ei_m.group(1)

            # If no CPE in version line, check subsequent script output lines for CPE
            if not cpe:
                for j in range(i+1, min(i+10, len(lines))):
                    sl = lines[j].strip()
                    if sl and not sl.startswith('|') and not sl.startswith('#'):
                        break
                    cpe_sm = re.search(r'(cpe:/[^\s|]+)', sl, re.IGNORECASE)
                    if cpe_sm:
                        cpe = cpe_sm.group(1)
                        break

            # Banner: look for banner script output on next lines
            banner = None
            for j in range(i+1, min(i+8, len(lines))):
                sl = lines[j].strip()
                if sl and not sl.startswith('|') and not sl.startswith('#'):
                    break
                bm = re.match(r'\|\s*(?:banner|http-server-header):\s*(.+)', sl, re.I)
                if bm:
                    banner = bm.group(1).strip()
                    break

            version = rest[:120] if rest else None

            data['ports'].append({
                'port': port,
                'protocol': proto,
                'service': svc,
                'version': version,
                'state': 'open',
                'extra_info': extra_info or banner,
                'cpe': cpe,
                'banner': banner,
            })

    # Port range from the ports we found
    if port_list_for_range:
        port_list_for_range.sort()
        if len(port_list_for_range) == 1:
            data['port_range'] = str(port_list_for_range[0])
        else:
            data['port_range'] = f"{port_list_for_range[0]} – {port_list_for_range[-1]}"
        data['ports_scanned'] = len(port_list_for_range)

    # Also look for "Scanning X ports" hint
    sp_m = re.search(r'Scanning (\d+) ports', nmap_output, re.IGNORECASE)
    if sp_m:
        data['ports_scanned'] = int(sp_m.group(1))

    # CVE extraction (CVE-YYYY-NNNNN+)
    for cve in set(re.findall(r'(CVE-\d{4}-\d{4,})', nmap_output, re.IGNORECASE)):
        data['vulnerabilities'].append({
            'name': cve,
            'description': f'Vulnerability {cve} detected',
            'severity': 'High',
            'source': 'Nmap',
            'cve': cve.upper(),
            'port': None
        })

    # Script vulnerabilities
    vuln_patterns = [
        (r'\|\s+VULNERABLE:\s*\n\|\s+(.+?)(?:\n\|(?!\s+)|$)', 'High'),
        (r'ssl-cert:.+?VULNERABLE', 'Medium'),
        (r'ssl-poodle:.+?VULNERABLE', 'High'),
        (r'ssl-dh-params:.+?WEAK', 'Medium'),
        (r'http-csrf:.+?Vulnerable', 'Medium'),
        (r'http-vuln.+?VULNERABLE', 'High'),
    ]
    for pattern, severity in vuln_patterns:
        for m in re.finditer(pattern, nmap_output, re.MULTILINE | re.DOTALL):
            desc = m.group(0).strip()[:200].replace('|', '').replace('\n', ' ')
            name = 'Security Issue'
            nm = re.search(r'(\w+(?:-\w+)*)', desc)
            if nm:
                name = nm.group(1)
            data['vulnerabilities'].append({
                'name': name,
                'description': desc,
                'severity': severity,
                'source': 'Nmap Script',
                'cve': None,
                'port': None
            })

    return data


def parse_nikto_output(nikto_output):
    """
    Parse nikto output to extract vulnerabilities and scan metadata.
    Returns a dict: {
        'findings': [...],        # list of vuln dicts
        'web_server': str|None,   # e.g. "Apache/2.4.41 (Ubuntu)"
        'scan_duration': str|None,# e.g. "12s"
        'target': str|None,
        'port': int|None,
    }
    For backwards compat, callers that expect a list should use result['findings'].
    """
    result_meta = {
        'findings': [],
        'web_server': None,
        'scan_duration': None,
        'target': None,
        'port': None,
    }
    if not nikto_output or len(nikto_output.strip()) < 10:
        return result_meta

    # Extract metadata from header lines
    start_time = None
    end_time = None
    for line in nikto_output.split('\n'):
        ls = line.strip()
        # "+ Server: Apache/2.4.41 (Ubuntu)"
        if ls.startswith('+ Server:') or ls.startswith('- Server:'):
            srv = ls.split(':', 1)[1].strip()
            if srv and result_meta['web_server'] is None:
                result_meta['web_server'] = srv
        # "+ Target IP: ..."
        elif 'Target IP:' in ls or 'Target Hostname:' in ls:
            m = re.search(r'(?:Target IP|Target Hostname):\s*(.+)', ls)
            if m and not result_meta['target']:
                result_meta['target'] = m.group(1).strip()
        # "+ Target Port: 80"
        elif 'Target Port:' in ls:
            m = re.search(r'Target Port:\s*(\d+)', ls)
            if m:
                result_meta['port'] = int(m.group(1))
        # "+ Start Time: 2026-03-26 05:28:25 (GMT+5.5)"
        elif 'Start Time:' in ls:
            m = re.search(r'Start Time:\s*(.+)', ls)
            if m:
                try:
                    from datetime import datetime as _dt
                    ts = re.sub(r'\s*\(.*\)\s*$', '', m.group(1).strip())
                    start_time = _dt.strptime(ts, '%Y-%m-%d %H:%M:%S')
                except Exception:
                    pass
        # "+ End Time: 2026-03-26 05:28:37 (GMT+5.5)"
        elif 'End Time:' in ls:
            m = re.search(r'End Time:\s*(.+)', ls)
            if m:
                try:
                    from datetime import datetime as _dt
                    ts = re.sub(r'\s*\(.*\)\s*$', '', m.group(1).strip())
                    end_time = _dt.strptime(ts, '%Y-%m-%d %H:%M:%S')
                except Exception:
                    pass

    if start_time and end_time:
        secs = int((end_time - start_time).total_seconds())
        result_meta['scan_duration'] = f"{secs}s"

    skip_patterns = [
        'target ip:', 'target hostname:', 'target port:', 'start time:', 'end time:',
        'server:', 'requests:', '----------', 'nikto v', 'no web server found',
        'testing:', 'retrieved', 'ssl info:'
    ]

    for line in nikto_output.split('\n'):
        line = line.strip()
        if not line.startswith('+ '):
            continue
        finding = line[2:].strip()
        if len(finding) < 10:
            continue
        if any(skip in finding.lower() for skip in skip_patterns):
            continue

        severity = 'Low'
        fl = finding.lower()
        high_kw = ['vulnerability', 'exploit', 'injection', 'sql', 'authentication bypass',
                   'remote code', 'arbitrary file', 'shell', 'xss', 'csrf', 'command injection',
                   'lfi', 'rfi', 'directory traversal', 'path traversal', 'cve-']
        med_kw = ['outdated', 'deprecated', 'disclosure', 'information leak', 'misconfiguration',
                  'weak', 'insecure', 'cleartext', 'unencrypted', 'default', 'exposed']
        if any(k in fl for k in high_kw):
            severity = 'High'
        elif any(k in fl for k in med_kw):
            severity = 'Medium'

        cve_match = re.search(r'(CVE-\d{4}-\d{4,})', finding, re.IGNORECASE)
        vuln_name = cve_match.group(1) if cve_match else 'Web Vulnerability'
        if 'header' in fl:
            vuln_name = 'Missing Security Header'
        elif 'directory' in fl and ('index' in fl or 'list' in fl):
            vuln_name = 'Directory Listing'
        elif 'outdated' in fl:
            vuln_name = 'Outdated Software'
        elif 'method' in fl:
            vuln_name = 'Unsafe HTTP Methods'
        elif 'ssl' in fl or 'tls' in fl:
            vuln_name = 'SSL/TLS Configuration Issue'

        result_meta['findings'].append({
            'source': 'Nikto',
            'name': vuln_name,
            'description': finding,
            'severity': severity,
            'cve': cve_match.group(1).upper() if cve_match else None,
            'port': result_meta['port'] or 80,
        })

    return result_meta


def get_remediation_steps(vulnerability):
    """Get remediation steps based on vulnerability type."""
    db = {
        'ssl-ccs-injection': {
            'steps': [
                'Update OpenSSL to the latest stable version',
                'Disable SSLv3 and use only TLS 1.2 or higher',
                'Review and update SSL/TLS configuration',
                'Implement Perfect Forward Secrecy (PFS)'
            ],
            'commands': [
                'sudo apt update && sudo apt upgrade openssl',
                'Edit /etc/ssl/openssl.cnf to disable SSLv3',
                'sudo systemctl restart apache2',
                'Test with: nmap --script ssl-enum-ciphers -p 443 <target>'
            ]
        },
        'outdated_apache': {
            'steps': [
                'Update Apache to the latest stable version',
                'Enable automatic security updates',
                'Review server configuration for security best practices',
                'Implement security headers'
            ],
            'commands': [
                'sudo apt update && sudo apt upgrade apache2',
                'sudo apt install unattended-upgrades',
                'Add security headers to /etc/apache2/conf-available/security.conf',
                'sudo systemctl restart apache2'
            ]
        },
        'missing_headers': {
            'steps': [
                'Configure X-Frame-Options header to prevent clickjacking',
                'Enable X-XSS-Protection header',
                'Set X-Content-Type-Options header',
                'Implement Content Security Policy (CSP)'
            ],
            'commands': [
                'Add to Apache config: Header set X-Frame-Options "SAMEORIGIN"',
                'Add: Header set X-XSS-Protection "1; mode=block"',
                'Add: Header set X-Content-Type-Options "nosniff"',
                'sudo systemctl reload apache2'
            ]
        },
        'directory_indexing': {
            'steps': [
                'Disable directory indexing in web server configuration',
                'Create index files for all directories',
                'Review file permissions',
                'Implement access controls'
            ],
            'commands': [
                'Edit httpd.conf or apache2.conf: Options -Indexes',
                'Create index.html files in all directories',
                'sudo chmod 750 /var/www/html/*',
                'sudo systemctl restart apache2'
            ]
        },
        'default': {
            'steps': [
                'Review the security advisory for the specific vulnerability',
                'Apply vendor-recommended patches immediately',
                'Test the fix in a staging environment',
                'Monitor for similar vulnerabilities'
            ],
            'commands': [
                'Check vendor security bulletins',
                'sudo apt update && sudo apt upgrade',
                'Review system logs for exploitation attempts',
                'Implement monitoring and alerting'
            ]
        }
    }
    desc = vulnerability.get('description', '').lower()
    if 'ssl' in desc or 'tls' in desc:
        return db['ssl-ccs-injection']
    if 'apache' in desc and 'outdated' in desc:
        return db['outdated_apache']
    if 'x-frame-options' in desc or 'x-xss-protection' in desc or 'header' in desc:
        return db['missing_headers']
    if 'directory indexing' in desc or 'directory listing' in desc:
        return db['directory_indexing']
    return db['default']


def get_installed_software_wmi(ip, username, password, domain=''):
    """
    Query installed software on a Windows host via WMI (DCOM/RPC).
    Uses impacket if available, otherwise falls back to Win32 wmic subprocess (Windows-only).

    Returns a list of software_inventory dicts compatible with the bulk scan result schema,
    or {'error': str} on failure.

    Credential format:
      - username: 'DOMAIN\\user' or plain 'user'
      - password: plain-text password
      - domain:   optional NetBIOS domain (can also be embedded in username)

    Note: Win32reg_AddRemovePrograms (registry) is tried first — it is faster and
    non-invasive. Win32_Product triggers Windows Installer reconfiguration on every query.
    """
    # ── Prefer impacket (cross-platform, works over TCP 135 + dynamic RPC) ──
    try:
        from impacket.dcerpc.v5.dcom import wmi as _wmi
        from impacket.dcerpc.v5.dcomrt import DCOMConnection

        _domain = domain or ''
        _user   = username or ''
        if not _domain and '\\' in _user:
            _domain, _user = _user.split('\\', 1)

        dcom = DCOMConnection(
            ip,
            username=_user,
            password=password,
            domain=_domain,
            oxidResolver=True,
        )
        iInterface       = dcom.CoCreateInstanceEx(_wmi.CLSID_WbemLevel1Login, _wmi.IID_IWbemLevel1Login)
        iWbemLevel1Login = _wmi.IWbemLevel1Login(iInterface)
        iWbemServices    = iWbemLevel1Login.NTLMLogin('//./root/cimv2', None, None)

        # Registry-based query first (fast, no side-effects), fallback to Win32_Product
        queries = [
            'SELECT Name, Version, Vendor, InstallDate FROM Win32reg_AddRemovePrograms',
            'SELECT Name, Version, Vendor, InstallDate FROM Win32_Product',
        ]
        software_list = []
        for query in queries:
            try:
                iEnum = iWbemServices.ExecQuery(query)
                while True:
                    try:
                        pEnum  = iEnum.Next(0xffffffff, 1)[0]
                        record = pEnum.getProperties()
                        name    = (record.get('Name')        or {}).get('value') or ''
                        version = (record.get('Version')     or {}).get('value') or 'Unknown'
                        vendor  = (record.get('Vendor')      or {}).get('value') or 'Unknown'
                        inst_dt = (record.get('InstallDate') or {}).get('value') or ''
                        if not name:
                            continue
                        software_list.append({
                            'display_name':     name,
                            'software_key':     name.lower().replace(' ', '_'),
                            'detected_version': str(version),
                            'vendor':           str(vendor),
                            'install_date':     str(inst_dt),
                            'category':         'Installed Software',
                            'latest_stable':    'Unknown',
                            'status':           'unknown',
                            'known_cves':       [],
                            'upgrade_url':      '',
                            'source':           'WMI',
                            'port':             None,
                        })
                    except Exception:
                        break
                if software_list:
                    break  # first working query succeeded
            except Exception:
                continue

        dcom.disconnect()
        return software_list

    except ImportError:
        pass  # impacket not installed

    except Exception as _wmi_exc:
        _msg = str(_wmi_exc)
        # DCE RPC 0x721 = WMI service unreachable (firewall blocking port 135)
        if '00000721' in _msg or 'RPC_S_SERVER_UNAVAILABLE' in _msg.upper():
            return {'error': 'WMI blocked on target — enable WMI in Windows Firewall on the target machine (port 135)'}
        if 'STATUS_LOGON_FAILURE' in _msg or 'logon_failure' in _msg.lower():
            return {'error': 'WMI auth failed — wrong username or password'}
        if 'STATUS_ACCESS_DENIED' in _msg or 'access_denied' in _msg.lower():
            return {'error': 'WMI access denied — account needs admin rights on target'}
        return {'error': f'WMI connection error: {_msg}'} 

    # ── Fallback: wmic subprocess (only works on Windows host) ────────────────
    import platform as _platform
    if _platform.system().lower() != 'windows':
        return {'error': 'impacket not installed and wmic is only available on Windows.'}

    try:
        _user_arg = f'{domain}\\{username}' if domain else username
        cmd = [
            'wmic',
            f'/node:{ip}',
            f'/user:{_user_arg}',
            f'/password:{password}',
            'product', 'get', 'Name,Version,Vendor',
            '/format:csv',
        ]
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        output = proc.stdout or ''
        software_list = []
        lines = [l for l in output.splitlines() if l.strip() and ',' in l]
        if lines:
            headers = [h.strip() for h in lines[0].split(',')]
            for line in lines[1:]:
                vals = [v.strip() for v in line.split(',')]
                row  = dict(zip(headers, vals))
                name = row.get('Name', '').strip()
                if not name:
                    continue
                software_list.append({
                    'display_name':     name,
                    'software_key':     name.lower().replace(' ', '_'),
                    'detected_version': row.get('Version', 'Unknown').strip(),
                    'vendor':           row.get('Vendor', 'Unknown').strip(),
                    'install_date':     '',
                    'category':         'Installed Software',
                    'latest_stable':    'Unknown',
                    'status':           'unknown',
                    'known_cves':       [],
                    'upgrade_url':      '',
                    'source':           'WMI',
                    'port':             None,
                })
        return software_list
    except subprocess.TimeoutExpired:
        return {'error': f'WMI query timed out for {ip}'}
    except FileNotFoundError:
        return {'error': 'wmic not found — install impacket or run scanner on Windows.'}
    except Exception as e:
        return {'error': str(e)}


def scan_single_ip(ip, modules=None, stopped_callback=None, port_depth='full',
                   wmi_username='', wmi_password='', wmi_domain=''):
    """
    Perform full scan on a single IP.
    Returns dict with all scan detail fields needed by the frontend tabs:
      ip, status, hostname, os, ports (with cpe/extra_info/banner),
      vulnerabilities, severity, scanned_at,
      ping_detail, latency, mac_address,
      scan_duration, ports_scanned, port_range, scan_technique,
      nikto (dict with web_server, scan_duration, findings)
    stopped_callback: callable that returns True if scan should stop
    wmi_username / wmi_password / wmi_domain: WMI credentials for Windows software inventory
    """
    modules = modules or ['ping', 'ports', 'services']
    result = {
        'ip': ip,
        'status': 'offline',
        'hostname': 'Unknown',
        'os': 'Unknown',
        'ports': [],
        'open_ports': [],
        'services': [],
        'vulnerabilities': [],
        'severity': 'none',
        'risk': 'None',
        'ping': False,
        'scanned_at': datetime.now().isoformat(),
        # Detailed fields for frontend tabs
        'ping_detail': None,
        'latency': None,
        'mac_address': None,
        'scan_duration': None,
        'ports_scanned': None,
        'port_range': None,
        'scan_technique': 'TCP SYN',
        'nikto': None,
    }

    try:
        # ── ARP/Hostname Discovery (fast pre-scan) ───────────────────────────
        # Run nmap -sn WITHOUT -Pn so ARP is used on LAN → gets MAC address.
        # Also does reverse DNS + NetBIOS for real hostname.
        discovery = run_nmap_discovery(ip)
        if discovery.get('mac_address'):
            result['mac_address'] = discovery['mac_address']
        if discovery.get('hostname'):
            result['hostname'] = discovery['hostname']
        if discovery.get('latency_ms'):
            result['latency'] = discovery['latency_ms']

        # ── Ping ─────────────────────────────────────────────────────────────
        if 'ping' in modules:
            ping_detail = run_ping_detailed(ip)
            ping_ok = ping_detail['alive']
            result['ping'] = ping_ok
            result['ping_detail'] = ping_detail
            # Prefer detailed ping latency; fallback to discovery latency
            if ping_detail.get('latency') is not None:
                result['latency'] = ping_detail['latency']
            elif result['latency'] and ping_detail:
                ping_detail['latency'] = result['latency']
                ping_detail['avg_rtt'] = f"{result['latency']} ms"
        else:
            ping_ok = True  # Assume reachable if we skip ping

        # ── Nmap ──────────────────────────────────────────────────────────────
        nmap_modules = []
        if 'ports' in modules or 'services' in modules:
            nmap_modules.append('ports')
        if 'services' in modules:
            nmap_modules.append('services')
        if 'os' in modules:
            nmap_modules.append('os')
        if 'vuln' in modules:
            nmap_modules.append('vuln')
        if not nmap_modules:
            nmap_modules = ['ports', 'services']

        if stopped_callback and stopped_callback():
            return result

        nmap_output = run_nmap_scan(ip, nmap_modules, port_depth=port_depth, stopped_callback=stopped_callback)
        nmap_data = parse_nmap_output(nmap_output)

        result['hostname'] = nmap_data['hostname']
        result['os'] = nmap_data['os']
        result['ports'] = nmap_data['ports']
        result['open_ports'] = [str(p['port']) for p in nmap_data['ports']]
        result['services'] = [f"{p['port']}/{p['service']}" for p in nmap_data['ports']]
        result['vulnerabilities'] = list(nmap_data['vulnerabilities'])

        # Hostname: use nmap_data if it found one, otherwise keep discovery result
        if nmap_data['hostname'] != 'Unknown':
            result['hostname'] = nmap_data['hostname']
        # else result['hostname'] already set from discovery above

        # MAC address: nmap port-scan output may also contain MAC (if run as root on LAN)
        # Use it only if discovery didn't already get one
        if nmap_data.get('mac_address') and not result.get('mac_address'):
            result['mac_address'] = nmap_data['mac_address']
        result['scan_duration'] = nmap_data.get('scan_duration')
        result['ports_scanned'] = nmap_data.get('ports_scanned')
        result['port_range']    = nmap_data.get('port_range')
        result['scan_technique']= nmap_data.get('scan_technique', 'TCP SYN')

        # If nmap found a latency hint and we don't have one from ping yet, use it
        if not result['latency']:
            lat_m = re.search(r'Host is up \(([\d.]+)s latency\)', nmap_output)
            if lat_m:
                result['latency'] = round(float(lat_m.group(1)) * 1000, 2)  # convert to ms
                if result['ping_detail']:
                    result['ping_detail']['latency'] = result['latency']
                    result['ping_detail']['avg_rtt'] = f"{result['latency']} ms"

        # OpenVAS / multi-layer OS detection
        result['_nmap_raw'] = nmap_output
        if _OPENVAS_OS_ENABLED:
            result = enrich_result_with_os(result, openvas_config=_OPENVAS_CONFIG)
            log_output(
                f"[OS] {ip} -> {result.get('os', 'Unknown')} "
                f"(source: {result.get('os_detail', {}).get('source', '?')}, "
                f"confidence: {result.get('os_detail', {}).get('confidence', 0)}%)"
            )
        result.pop('_nmap_raw', None)

        # ── Nikto (only if web ports open) ────────────────────────────────────
        web_ports = [p['port'] for p in nmap_data['ports'] if p['port'] in ['80', '443', '8080', '8443']]
        if 'nikto' in modules and web_ports:
            if not (stopped_callback and stopped_callback()):
                nikto_output = run_nikto_scan(ip, web_ports=web_ports)
                nikto_meta = parse_nikto_output(nikto_output)
                nikto_vulns = nikto_meta['findings']
                result['vulnerabilities'].extend(nikto_vulns)
                # Store full nikto detail for the Nikto tab
                result['nikto'] = {
                    'web_server':    nikto_meta.get('web_server'),
                    'scan_duration': nikto_meta.get('scan_duration'),
                    'target':        nikto_meta.get('target'),
                    'port':          nikto_meta.get('port'),
                    'findings':      nikto_vulns,
                    'has_http':      True,
                }
        elif 'nikto' in modules and not web_ports:
            # Nikto was requested but no HTTP ports — record that for the tab
            result['nikto'] = {
                'web_server':    None,
                'scan_duration': None,
                'findings':      [],
                'has_http':      False,
            }

        # Nuclei (template-based vuln scan, only if web ports open)
        if 'nuclei' in modules and web_ports:
            if not (stopped_callback and stopped_callback()):
                log_output(f"[NUCLEI] Scanning {ip} (web ports: {web_ports})")
                nuclei_vulns = run_nuclei_scan(ip, web_ports=web_ports)
                result['vulnerabilities'].extend(nuclei_vulns)

        # ── Software Inventory & Outdated-Version Vulnerability Detection ──────
        # Analyses detected service banners to identify installed software,
        # check versions against the known-version database, and inject
        # vulnerability entries (with full remediation) for outdated/EOL software.
        result = enrich_scan_result_with_software(result)

        # ── WMI Software Inventory (Windows hosts) ────────────────────────────
        # Called when the 'wmi' module is selected and credentials are supplied.
        # Merges the WMI-discovered software list into result['software_inventory'],
        # de-duplicating by display_name so banner-detected + WMI entries coexist.
        if 'wmi' in modules:
            # Use credentials passed directly as function parameters (preferred),
            # falling back to any pre-set keys in the result dict for legacy callers.
            _wmi_user   = wmi_username or result.get('_wmi_username', '')
            _wmi_pass   = wmi_password if wmi_password is not None else result.get('_wmi_password', '')
            _wmi_domain = wmi_domain   or result.get('_wmi_domain', '')
            # Allow blank password — some machines have no password set
            if _wmi_user:
                log_output(f"[WMI] Querying installed software on {ip} (user: {_wmi_user})")
                try:
                    wmi_result = get_installed_software_wmi(ip, _wmi_user, _wmi_pass, _wmi_domain)
                except Exception as _wmi_exc:
                    wmi_result = {'error': str(_wmi_exc)}
                if isinstance(wmi_result, list) and wmi_result:
                    existing_names = {
                        sw.get('display_name', '').lower()
                        for sw in result.get('software_inventory', [])
                    }
                    merged = list(result.get('software_inventory', []))
                    for sw in wmi_result:
                        if sw.get('display_name', '').lower() not in existing_names:
                            merged.append(sw)
                            existing_names.add(sw.get('display_name', '').lower())
                    result['software_inventory'] = merged
                    log_output(f"[WMI] Found {len(wmi_result)} installed package(s) on {ip}")
                elif isinstance(wmi_result, dict) and wmi_result.get('error'):
                    log_output(f"[WMI] Error on {ip}: {wmi_result['error']}")
                    result['wmi_error'] = wmi_result['error']
            else:
                log_output(f"[WMI] Skipping {ip} — no WMI username provided")


        vulns = result['vulnerabilities']
        high_sev = sum(1 for v in vulns if v.get('severity') == 'High')
        crit_sev = sum(1 for v in vulns if v.get('severity') == 'Critical')
        risky_ports = {'21', '23', '3389', '445', '3306'}
        has_risky = any(p in risky_ports for p in result['open_ports'])

        if crit_sev or (high_sev >= 2) or has_risky:
            result['severity'] = 'critical'
            result['risk'] = 'High'
        elif high_sev or has_risky:
            result['severity'] = 'high'
            result['risk'] = 'High'
        elif vulns:
            result['severity'] = 'medium'
            result['risk'] = 'Medium'
        elif result['open_ports']:
            result['severity'] = 'low'
            result['risk'] = 'Medium'
        elif ping_ok:
            result['severity'] = 'none'
            result['risk'] = 'Low'
        else:
            result['severity'] = 'none'
            result['risk'] = 'None'

        # Online if we got any useful data
        # Note: WSL ping often fails for LAN IPs due to network isolation,
        # so also check for ports, hostname, or OS data from nmap.
        nmap_found_something = bool(result['ports'] or
                                    (result.get('hostname') and result['hostname'] != 'Unknown') or
                                    (result.get('os') and result['os'] != 'Unknown'))
        if ping_ok or nmap_found_something:
            result['status'] = 'online'
        else:
            result['status'] = 'offline'

    except Exception as e:
        result['status'] = 'error'
        result['error'] = str(e)
        result['severity'] = 'none'

    return result


def create_bulk_excel_report(scan_results, filepath=None):
    """Create Excel report for bulk scan results. Returns filepath."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    title_font = Font(name='Arial', size=20, bold=True, color='1F4E78')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    border_thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Summary sheet
    ws = wb.create_sheet('Summary')
    ws['A1'] = 'Bulk IP Security Scan Report'
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
    ws.merge_cells('A2:F2')

    headers = ['IP Address', 'Hostname', 'OS', 'Open Ports', 'Vulnerabilities', 'Severity']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col)
        c.value = h
        c.font = header_font
        c.fill = header_fill
        c.border = border_thin
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for idx, r in enumerate(scan_results):
        row = 5 + idx
        fill = PatternFill(start_color='F2F2F2' if idx % 2 == 0 else 'FFFFFF', end_color='F2F2F2' if idx % 2 == 0 else 'FFFFFF', fill_type='solid')
        ws.cell(row=row, column=1, value=r.get('ip', '')).border = border_thin
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=2, value=r.get('hostname', 'Unknown')).border = border_thin
        ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=3, value=r.get('os', 'Unknown')).border = border_thin
        ws.cell(row=row, column=3).fill = fill
        ws.cell(row=row, column=4, value=len(r.get('ports', []))).border = border_thin
        ws.cell(row=row, column=4).fill = fill
        ws.cell(row=row, column=5, value=len(r.get('vulnerabilities', []))).border = border_thin
        ws.cell(row=row, column=5).fill = fill
        sev_cell = ws.cell(row=row, column=6, value=(r.get('severity') or 'none').upper())
        sev_cell.border = border_thin
        sev_cell.fill = fill
        sev_cell.alignment = Alignment(horizontal='center')

    for i in range(1, 7):
        ws.column_dimensions[get_column_letter(i)].width = [18, 30, 35, 12, 15, 12][i - 1]

    # Vulnerabilities sheet
    wv = wb.create_sheet('Vulnerabilities')
    wv['A1'] = 'Detected Vulnerabilities'
    wv['A1'].font = Font(name='Arial', size=14, bold=True, color='C00000')
    wv.merge_cells('A1:E1')
    v_headers = ['IP Address', 'Vulnerability', 'Severity', 'Source', 'Description']
    for col, h in enumerate(v_headers, 1):
        c = wv.cell(row=3, column=col)
        c.value = h
        c.font = header_font
        c.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        c.border = border_thin
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    row = 4
    for r in scan_results:
        vulns = r.get('vulnerabilities', [])
        if not vulns:
            wv.cell(row=row, column=1, value=r.get('ip', '')).border = border_thin
            wv.cell(row=row, column=2, value='No vulnerabilities detected').border = border_thin
            wv.merge_cells(f'B{row}:E{row}')
            row += 1
        else:
            for v in vulns:
                wv.cell(row=row, column=1, value=r.get('ip', '')).border = border_thin
                wv.cell(row=row, column=2, value=v.get('name', 'Unknown')).border = border_thin
                wv.cell(row=row, column=3, value=v.get('severity', 'Medium')).border = border_thin
                wv.cell(row=row, column=4, value=v.get('source', 'Nmap')).border = border_thin
                wv.cell(row=row, column=5, value=v.get('description', '')[:500]).border = border_thin
                wv.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
                row += 1

    for i in range(1, 6):
        wv.column_dimensions[get_column_letter(i)].width = [18, 30, 12, 12, 70][i - 1]

    # ── Software Inventory sheet ───────────────────────────────────────────
    wi = wb.create_sheet('Software Inventory')
    wi['A1'] = 'Detected Software & Version Status'
    wi['A1'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
    wi.merge_cells('A1:H1')

    sw_headers = ['IP Address', 'Software', 'Category', 'Detected Version',
                  'Latest Stable', 'Status', 'Known CVEs', 'Upgrade URL']
    status_colors = {
        'eol': 'C00000',       # dark red
        'critical': 'FF0000',  # red
        'outdated': 'FF6600',  # orange
        'current': '00B050',   # green
        'unknown': '808080',   # grey
    }
    for col, h in enumerate(sw_headers, 1):
        c = wi.cell(row=3, column=col)
        c.value = h
        c.font = header_font
        c.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        c.border = border_thin
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    sw_row = 4
    for r in scan_results:
        sw_inventory = r.get('software_inventory', [])
        if not sw_inventory:
            # Still show the IP with "No software detected"
            wi.cell(row=sw_row, column=1, value=r.get('ip', '')).border = border_thin
            wi.cell(row=sw_row, column=2, value='No software detected').border = border_thin
            wi.merge_cells(f'B{sw_row}:H{sw_row}')
            sw_row += 1
        else:
            for sw in sw_inventory:
                status = sw.get('status', 'unknown')
                color = status_colors.get(status, 'FFFFFF')
                row_fill = PatternFill(start_color=color + '33' if len(color) == 6 else 'FFFFFF',
                                       end_color=color + '33' if len(color) == 6 else 'FFFFFF',
                                       fill_type='solid')
                wi.cell(row=sw_row, column=1, value=r.get('ip', '')).border = border_thin
                wi.cell(row=sw_row, column=2, value=sw.get('display_name', '')).border = border_thin
                wi.cell(row=sw_row, column=3, value=sw.get('category', '')).border = border_thin
                wi.cell(row=sw_row, column=4, value=sw.get('detected_version', 'Unknown')).border = border_thin
                wi.cell(row=sw_row, column=5, value=sw.get('latest_stable', 'Unknown')).border = border_thin
                status_cell = wi.cell(row=sw_row, column=6, value=status.upper())
                status_cell.border = border_thin
                status_cell.font = Font(name='Arial', bold=True, color=color)
                status_cell.alignment = Alignment(horizontal='center')
                cves = sw.get('known_cves', [])
                wi.cell(row=sw_row, column=7, value=', '.join(cves[:3])).border = border_thin
                wi.cell(row=sw_row, column=8, value=sw.get('upgrade_url', '')).border = border_thin
                wi.cell(row=sw_row, column=8).alignment = Alignment(wrap_text=True)
                for col in range(1, 9):
                    wi.cell(row=sw_row, column=col).fill = row_fill
                sw_row += 1

    sw_col_widths = [18, 28, 18, 18, 16, 12, 40, 50]
    for i, w in enumerate(sw_col_widths, 1):
        wi.column_dimensions[get_column_letter(i)].width = w

    # ── Remediation sheet ─────────────────────────────────────────────────
    wr = wb.create_sheet('Remediations')
    wr['A1'] = 'Outdated Software Remediation Guide'
    wr['A1'].font = Font(name='Arial', size=14, bold=True, color='C00000')
    wr.merge_cells('A1:F1')

    rem_headers = ['IP Address', 'Software', 'Version', 'Status', 'Remediation Steps', 'Commands']
    for col, h in enumerate(rem_headers, 1):
        c = wr.cell(row=3, column=col)
        c.value = h
        c.font = header_font
        c.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        c.border = border_thin
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    rem_row = 4
    for r in scan_results:
        sw_inventory = r.get('software_inventory', [])
        for sw in sw_inventory:
            if sw.get('status') not in ('eol', 'critical', 'outdated'):
                continue
            rem = sw.get('remediation') or {}
            steps = rem.get('steps', [])
            commands = rem.get('commands', [])
            wr.cell(row=rem_row, column=1, value=r.get('ip', '')).border = border_thin
            wr.cell(row=rem_row, column=2, value=sw.get('display_name', '')).border = border_thin
            wr.cell(row=rem_row, column=3, value=sw.get('detected_version', '')).border = border_thin
            wr.cell(row=rem_row, column=4, value=(sw.get('status') or '').upper()).border = border_thin
            steps_text = '\n'.join(f'{i+1}. {s}' for i, s in enumerate(steps))
            steps_cell = wr.cell(row=rem_row, column=5, value=steps_text)
            steps_cell.border = border_thin
            steps_cell.alignment = Alignment(wrap_text=True, vertical='top')
            cmds_text = '\n'.join(f'$ {c}' for c in commands)
            cmd_cell = wr.cell(row=rem_row, column=6, value=cmds_text)
            cmd_cell.border = border_thin
            cmd_cell.alignment = Alignment(wrap_text=True, vertical='top')
            wr.row_dimensions[rem_row].height = max(60, len(steps) * 18)
            rem_row += 1

    rem_col_widths = [18, 24, 14, 12, 70, 55]
    for i, w in enumerate(rem_col_widths, 1):
        wr.column_dimensions[get_column_letter(i)].width = w

    if not filepath:
        filepath = f'{REPORTS_DIR}/bulk_scan_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(filepath)
    return filepath